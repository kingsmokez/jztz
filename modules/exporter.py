"""Export helpers — CSV (stdlib) + Excel (.xlsx via openpyxl).

Used by ``routes/export.py`` and available as a library for any
synchronous export (downloads, scheduled reports, etc.).

Public surface
--------------
* :func:`to_csv`     — returns ``bytes`` of UTF-8-with-BOM CSV
* :func:`to_xlsx`    — returns ``bytes`` of an ``.xlsx`` workbook
* :func:`pick_columns` — choose which columns to export (default: all)

Why UTF-8-with-BOM?
    Excel on Windows decodes ``\ufeff`` as a BOM, then autodetects
    UTF-8 correctly. Without it, Chinese characters look like garbage.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence

# OpenPyXL is the only third-party dep this module needs.  If the
# runtime is offline and it isn't installed, the XLSX helpers raise
# ``ExporterError`` so callers can fall back to CSV.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover — guarded by import test
    Workbook = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------
class ExporterError(RuntimeError):
    """Raised when an export cannot be produced."""


# ---------------------------------------------------------------------------
# Column helpers
# ---------------------------------------------------------------------------
_NUMERIC_RE = re.compile(r"^-?\d+(\.\d+)?$")
_NUMERIC_KEYS = {
    "score", "pe", "pb", "roe", "change_pct", "change",
    "price", "close", "open", "high", "low", "volume",
    "turnover_rate", "amount", "market_cap", "circulation_cap",
    "pe_ttm", "pb_mrq", "ps_ttm", "debt_ratio", "net_margin",
    "gross_margin", "dividend_yield", "total_score", "weight",
    "quantity", "cost", "value", "profit", "profit_pct", "shares",
}


def is_numeric_key(key: str) -> bool:
    """Return True if a column key looks like a numeric field."""
    if key in _NUMERIC_KEYS:
        return True
    if key.endswith("_pct") or key.endswith("_rate") or key.endswith("_ratio"):
        return True
    return False


def pick_columns(
    rows: Sequence[Dict[str, Any]],
    prefer: Optional[Sequence[str]] = None,
) -> List[str]:
    """Return a stable column ordering.

    * If ``prefer`` is given AND every preferred column is present in
      ``rows``, returns ``prefer``.
    * Otherwise returns the union of all keys, in first-seen order.
    """
    if prefer and rows:
        all_keys = {k for r in rows for k in r}
        if all(p in all_keys for p in prefer):
            return list(prefer)
    seen: List[str] = []
    for r in rows:
        for k in r:
            if k not in seen:
                seen.append(k)
    return seen


def _stringify(value: Any) -> str:
    """Best-effort ``str()`` for any cell value.

    Always returns a ``str`` — callers can rely on it for display,
    length measurement, or concatenation. Numerics are formatted
    with ``str()`` to keep their Python repr (e.g. ``5.2``,
    ``0.0001``); boolean is rendered as ``true``/``false`` so
    it doesn't get mistaken for the integer ``1``/``0``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (list, tuple)):
        return ",".join(str(x) for x in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value == "":
        return True
    if isinstance(value, (list, tuple, dict)) and len(value) == 0:
        return True
    return False


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value: str) -> str:
    """Prefix values that Excel/Sheets would treat as a formula.

    A leading ``=``, ``+``, ``-``, ``@`` (or tab/CR) in a CSV cell is
    interpreted as a formula when the file is opened in Excel. An
    attacker who can plant such a value (e.g. a stock name like
    ``=cmd|'/c calc'!A0``) could trigger arbitrary code execution on
    the viewer's machine. Prefixing with a single quote disables
    formula parsing while keeping the original text visible.
    """
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def to_csv(
    rows: Iterable[Dict[str, Any]],
    columns: Optional[Sequence[str]] = None,
    sheet_name: str = "data",
) -> bytes:
    """Encode ``rows`` as a UTF-8-with-BOM CSV.

    Parameters
    ----------
    rows
        Iterable of dicts (one per row).
    columns
        Explicit column order. If ``None``, derived from the first row
        using :func:`pick_columns`.
    sheet_name
        Ignored (CSV is single-sheet). Accepted for API symmetry.
    """
    rows = list(rows)
    if not rows:
        return b"\xef\xbb\xbf"

    if columns is None:
        columns = pick_columns(rows)

    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(columns)
    for r in rows:
        writer.writerow([_csv_safe(_stringify(r.get(c))) for c in columns])

    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def _require_openpyxl() -> None:
    if not _OPENPYXL_AVAILABLE:
        raise ExporterError(
            "openpyxl is not installed. Run `pip install openpyxl` "
            "or use format=csv."
        )


def to_xlsx(
    rows: Iterable[Dict[str, Any]],
    columns: Optional[Sequence[str]] = None,
    sheet_name: str = "data",
    title: Optional[str] = None,
) -> bytes:
    """Encode ``rows`` as an ``.xlsx`` workbook (single sheet).

    Numeric-looking columns are right-aligned, headers are bolded with
    a light-gray background, and a frozen header row makes the file
    Excel-friendly for human review.
    """
    _require_openpyxl()
    rows = list(rows)
    if columns is None:
        columns = pick_columns(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(sheet_name)

    # Optional banner row.
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        start_row = 2

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A5568")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    align_right = Alignment(horizontal="right")
    align_left = Alignment(horizontal="left")
    for r_idx, r in enumerate(rows, start_row + 1):
        for c_idx, key in enumerate(columns, 1):
            value = r.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=_stringify(value))
            cell.alignment = align_right if is_numeric_key(key) else align_left

    # Freeze header
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto column width
    for c_idx, key in enumerate(columns, 1):
        # Use max of header width, longest cell value (sample 200 rows).
        header_width = _display_width(str(key))
        body_width = 0
        for r in rows[:200]:
            v = r.get(key)
            body_width = max(body_width, _display_width(_stringify(v)))
        width = min(max(header_width, body_width) + 2, 50)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # Write to buffer
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\."""
    s = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "data"
    return s[:31]


def _display_width(value: Any) -> int:
    """Approximate display width: CJK chars count as 2."""
    s = value if isinstance(value, str) else _stringify(value)
    w = 0
    for ch in s:
        if ord(ch) > 0x2E80:  # rough CJK + wider block
            w += 2
        else:
            w += 1
    return w


# ---------------------------------------------------------------------------
# Data sources — wire exports to the live in-memory state
# ---------------------------------------------------------------------------
def collect_daily_quotes(session: Optional[str] = None) -> List[Dict[str, Any]]:
    """Return rows for the daily-pick endpoints.

    ``session`` may be ``"morning"``, ``"afternoon"``, or ``None`` (all).
    Each row is augmented with ``session`` and ``session_label`` fields
    so the user can see which pass it came from.
    """
    from routes.daily import DAILY_PICK_DATA  # late import (avoid cycles)

    sessions: List[tuple] = []
    if session in (None, "morning") and DAILY_PICK_DATA.get("morning"):
        sessions.append(("morning", DAILY_PICK_DATA["morning"]))
    if session in (None, "afternoon") and DAILY_PICK_DATA.get("afternoon"):
        sessions.append(("afternoon", DAILY_PICK_DATA["afternoon"]))

    rows: List[Dict[str, Any]] = []
    for key, block in sessions:
        for r in block.get("results", []) or []:
            row = dict(r)
            row["session"] = key
            row["session_label"] = block.get("session_type", key)
            row["pick_time"] = block.get("pick_time", "")
            rows.append(row)
    return rows


def collect_live_quotes(limit: int = 100) -> List[Dict[str, Any]]:
    """Return the most-recent live quote snapshot from the cache, if any.

    Falls back to an empty list when no snapshot is cached.
    """
    try:
        from modules.cache_manager import cache
        snap = cache.get("live_quotes_snapshot")
        if not snap:
            return []
        if isinstance(snap, list):
            return snap[:limit]
        if isinstance(snap, dict) and "data" in snap:
            return snap["data"][:limit]
    except Exception:
        return []
    return []


def collect_auction_quotes() -> List[Dict[str, Any]]:
    """Return auction-pick results, if cached."""
    try:
        from modules.cache_manager import cache
        snap = cache.get("auction_results")
        if not snap:
            return []
        if isinstance(snap, list):
            return snap
        if isinstance(snap, dict) and "results" in snap:
            return snap["results"]
    except Exception:
        return []
    return []


__all__ = [
    "ExporterError",
    "is_numeric_key",
    "pick_columns",
    "to_csv",
    "to_xlsx",
    "collect_daily_quotes",
    "collect_live_quotes",
    "collect_auction_quotes",
]
